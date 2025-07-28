import os

os.environ["OMP_NUM_THREADS"] = "2"
import sys
import pandas as pd
import numpy as np 
import requests
import pandas as pd
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill



from io import BytesIO
from glob import glob
from PIL import Image, ImageEnhance, ImageDraw, ImageFont

import streamlit as st

# Add current directory to import path to access local utils
sys.path.insert(0, ".")

# Import custom palette utilities (clustering, sorting, visualization, etc.)
from sophisticated_palette.utils import show_palette, model_dict, get_palette, \
    sort_func_dict, store_palette, display_matplotlib_code, display_plotly_code,\
     get_df_rgb, enhancement_range, plot_rgb_3d, plot_hsv_3d, print_praise

# Load all image file paths from the 'images' folder
gallery_files = glob(os.path.join(".", "images", "*"))
# Create a dictionary mapping image names (cleaned up) to file paths
gallery_dict = {
    os.path.splitext(os.path.basename(image_path))[0].replace("-", " "): image_path
    for image_path in gallery_files
}

# Display logo and app info in the sidebar    
st.image("logo.jpg")
st.sidebar.title("Sophisticated Palette 🎨")
st.sidebar.caption("Tell your data story with style.")
st.sidebar.markdown("Made by Rena Herman, Rivka Palace, Shira Cohen, Tehila Hakkakian, Miri Eisenberg")



st.sidebar.markdown("Checkout our github repo here: https://github.com/tehilahak/sophisticated_palette.git")
st.sidebar.markdown("---")

toggle = st.sidebar.checkbox("Toggle Update", value=True, help="Continuously update the pallete with every change in the app.")
click = st.sidebar.button("Find Palette", disabled=bool(toggle))
# Palette generation settings
st.sidebar.markdown("---")
st.sidebar.header("Settings")
palette_size = int(st.sidebar.number_input("palette size", min_value=1, max_value=20, value=5, step=1, help="Number of colors to infer from the image."))
sample_size = int(st.sidebar.number_input("sample size", min_value=5, max_value=3000, value=500, step=500, help="Number of sample pixels to pick from the image."))

# Image Enhancement
enhancement_categories = enhancement_range.keys()
enh_expander = st.sidebar.expander("Image Enhancements", expanded=False)
# Image enhancement sliders in sidebar (e.g., brightness, contrast, color)
with enh_expander:
    
    if st.button("reset"):
        # Reset all enhancements to default (1.0)
        for cat in enhancement_categories:
            if f"{cat}_enhancement" in st.session_state:
                st.session_state[f"{cat}_enhancement"] = 1.0
# Create sliders for each enhancement category
enhancement_factor_dict = {
    cat: enh_expander.slider(f"{cat} Enhancement", 
                            value=1., 
                            min_value=enhancement_range[cat][0], 
                            max_value=enhancement_range[cat][1], 
                            step=enhancement_range[cat][2],
                            key=f"{cat}_enhancement")
    for cat in enhancement_categories
}
# Tip for enhancement settings
enh_expander.info("**Try the following**\n\nColor Enhancements = 2.6\n\nContrast Enhancements = 1.1\n\nBrightness Enhancements = 1.1")

# Clustering Model 
model_name = st.sidebar.selectbox("machine learning model", model_dict.keys(), help="Machine Learning model to use for clustering pixels and colors together.")
sklearn_info = st.sidebar.empty()

sort_options = sorted(list(sort_func_dict.keys()) + [key + "_r" for key in sort_func_dict.keys() if key!="random"])
sort_func = st.sidebar.selectbox("palette sort function", options=sort_options, index=5)

# Random Number Seed
seed = int(st.sidebar.number_input("random seed", value=42, help="Seed used for all random samplings."))
np.random.seed(seed)
st.sidebar.markdown("---")


# =======
#   App
# =======

# provide options to either select an image form the gallery, upload one, or fetch from URL
gallery_tab, upload_tab, url_tab = st.tabs(["Gallery", "Upload", "Image URL"])
with gallery_tab:
    options = list(gallery_dict.keys())
    file_name = st.selectbox("Select Art", 
                            options=options, index=options.index("Mona Lisa (Leonardo da Vinci)"))
    file = gallery_dict[file_name]
 # Warn user if other input methods are active
    if st.session_state.get("file_uploader") is not None:
        st.warning("To use the Gallery, remove the uploaded image first.")
    if st.session_state.get("image_url") not in ["", None]:
        st.warning("To use the Gallery, remove the image URL first.")

    img = Image.open(file)

# Upload tab: user can upload their own image
with upload_tab:
    file = st.file_uploader("Upload Art", key="file_uploader")
    if file is not None:
        try:
            img = Image.open(file)
        except:
            st.error("The file you uploaded does not seem to be a valid image. Try uploading a png or jpg file.")
    if st.session_state.get("image_url") not in ["", None]:
        st.warning("To use the file uploader, remove the image URL first.")

# URL tab: image will be loaded from a given URL
with url_tab:
    url_text = st.empty()
    
    # FIXME: the button is a bit buggy, but it's worth fixing this later

    # url_reset = st.button("Clear URL", key="url_reset")
    # if url_reset and "image_url" in st.session_state:
    #     st.session_state["image_url"] = ""
    #     st.write(st.session_state["image_url"])

    url = url_text.text_input("Image URL", key="image_url")
    
    if url!="":
        try:
            response = requests.get(url)
            img = Image.open(BytesIO(response.content))
        except:
            st.error("The URL does not seem to be valid.")

# convert RGBA to RGB if necessary
n_dims = np.array(img).shape[-1]
if n_dims == 4:
    background = Image.new("RGB", img.size, (255, 255, 255))
    background.paste(img, mask=img.split()[3]) # 3 is the alpha channel
    img = background

# apply image enhancements
for cat in enhancement_categories:
    img = getattr(ImageEnhance, cat)(img)
    img = img.enhance(enhancement_factor_dict[cat])

# show the image
with st.expander("🖼  Artwork", expanded=True):
    st.image(img, use_column_width=True)


if click or toggle:
    
    df_rgb = get_df_rgb(img, sample_size)

    # (optional for later)
    # plot_rgb_3d(df_rgb) 
    # plot_hsv_3d(df_rgb) 

    # calculate the RGB palette and cache it to session_state
    st.session_state["palette_rgb"] = get_palette(df_rgb, model_name, palette_size, sort_func=sort_func)

    if "palette_rgb" in st.session_state:
        
        # store individual colors in session state
        store_palette(st.session_state["palette_rgb"])

        st.write("---")

        # sort the colors based on the selected option
        colors = {k: v for k, v in st.session_state.items() if k.startswith("col_")}
        sorted_colors = {k: colors[k] for k in sorted(colors, key=lambda k: int(k.split("_")[-1]))}
        
        # find the hex representation for matplotlib and plotly settings
        palette_hex = [color for color in sorted_colors.values()][:palette_size]
        # Show the palette and provide example code for using it
        with st.expander("Adopt this Palette", expanded=False):
            st.pyplot(show_palette(palette_hex))

            matplotlib_tab, plotly_tab = st.tabs(["matplotlib", "plotly"])

            with matplotlib_tab:
                display_matplotlib_code(palette_hex)

                # Set color cycle for matplotlib using the palette
                import matplotlib as mpl
                from cycler import cycler

                mpl.rcParams["axes.prop_cycle"] = cycler(color=palette_hex)
                import matplotlib.pyplot as plt


                # Create random example data
                x = np.arange(5)
                y_list = np.random.random((len(palette_hex), 5))+2
                df = pd.DataFrame(y_list).T

                area_tab, bar_tab = st.tabs(["area chart", "bar chart"])

                with area_tab:
                    fig_area , ax_area = plt.subplots()
                    df.plot(kind="area", ax=ax_area, backend="matplotlib", )  
                    st.header("Example Area Chart")
                    st.pyplot(fig_area)
    
                with bar_tab:
                    fig_bar , ax_bar = plt.subplots()
                    df.plot(kind="bar", ax=ax_bar, stacked=True, backend="matplotlib", )
                    st.header("Example Bar Chart")
                    st.pyplot(fig_bar)

                
            with plotly_tab:
                display_plotly_code(palette_hex)

                # Set up a custom Plotly template with the selected colors
                import plotly.io as pio
                import plotly.graph_objects as go
                pio.templates["sophisticated"] = go.layout.Template(
                    layout=go.Layout(
                    colorway=palette_hex
                    )
                )
                pio.templates.default = 'sophisticated'

                # Tabs to preview Plotly charts with the palette
                area_tab, bar_tab = st.tabs(["area chart", "bar chart"])

                # Display sample Plotly charts using the selected color palette
                with area_tab:
                    fig_area = df.plot(kind="area", backend="plotly", )
                    st.header("Example Area Chart")
                    st.plotly_chart(fig_area, use_container_width=True)
    
                with bar_tab:
                    fig_bar = df.plot(kind="bar", backend="plotly", barmode="stack")
                    st.header("Example Bar Chart")
                    st.plotly_chart(fig_bar, use_container_width=True)

       
else:
    st.info("👈  Click on 'Find Palette' ot turn on 'Toggle Update' to see the color palette.")

st.sidebar.success(print_praise())   
st.sidebar.write("---\n")
st.sidebar.caption("""You can check out the source code [here](https://github.com/syasini/sophisticated_palette).
                      The `matplotlib` and `plotly` code snippets have been borrowed from [here](https://matplotlib.org/stable/users/prev_whats_new/dflt_style_changes.html) and [here](https://stackoverflow.com/questions/63011674/plotly-how-to-change-the-default-color-pallete-in-plotly).""")
st.sidebar.write("---\n")

# --- Download Options ---
hex_str = "\n".join(palette_hex)
st.download_button("📄 Download Palette as TXT", hex_str, file_name="palette.txt")


# Create a new Excel file with a sheet named "Palette"
def create_colored_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Palette"

    # Write header
    ws.cell(row=1, column=1, value="Hex Code")

    # Write data and apply fill color
    for i, hex_code in enumerate(df["Hex Code"], start=2):
        cell = ws.cell(row=i, column=1, value=hex_code)
        fill = PatternFill(start_color=hex_code.lstrip('#').upper(),
                           end_color=hex_code.lstrip('#').upper(),
                           fill_type="solid")
        cell.fill = fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# Create DataFrame from your actual palette_hex
df_palette = pd.DataFrame({"Hex Code": palette_hex})

# Generate the Excel file bytes
excel_file = create_colored_excel(df_palette)

# Streamlit download button
st.download_button(
    label="📥 Download Palette as Excel (.xlsx)",
    data=excel_file,
    file_name="palette_colored.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
# Create a blank image to display color swatches side by side
def get_palette_image(colors, size=(300, 80)):
    num_colors = len(colors)
    swatch_width = size[0] // num_colors  # width of each color block
    swatch_height = 50

    img = Image.new("RGB", size, color="white")
    draw = ImageDraw.Draw(img)

    # Optional: load a better font (fallback to default if not found)
    try:
        font = ImageFont.truetype("arial.ttf", size=12)
    except IOError:
        font = ImageFont.load_default()

    for i, color in enumerate(colors):
        x0 = i * swatch_width
        x1 = (i + 1) * swatch_width
        # Draw swatch
        draw.rectangle([x0, 0, x1, swatch_height], fill=color)

        # Draw hex text centered under the swatch
        hex_code = color.upper()
        text_bbox = draw.textbbox((0, 0), hex_code, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_x = x0 + (swatch_width - text_width) // 2
        draw.text((text_x, swatch_height + 5), hex_code, fill="black", font=font)

    return img

# Generate and download the palette image as a PNG
img_palette = get_palette_image(palette_hex)
buf = BytesIO()
img_palette.save(buf, format="PNG")
st.download_button("🖼️ Download Palette Image", buf.getvalue(), file_name="palette.png", mime="image/png")

# Option to save the current palette with a custom name
with st.expander("Save this Palette", expanded=False):
    st.pyplot(show_palette(palette_hex))
    palette_name = st.text_input("Name your palette", value="My Palette")
     # Initialize storage if it doesn't exist yet
    if st.button("Save Palette"):
        if "saved_palettes" not in st.session_state:
            st.session_state["saved_palettes"] = []
        # Add current palette to saved list
        st.session_state["saved_palettes"].append({"name": palette_name, "colors": palette_hex})
        st.success(f"Palette '{palette_name}' saved!")

# Show saved palettes
if "saved_palettes" in st.session_state:
    st.header("🎨 Your Saved Palettes")
    for p in st.session_state["saved_palettes"]:
        st.write(f"**{p['name']}**: {p['colors']}")

